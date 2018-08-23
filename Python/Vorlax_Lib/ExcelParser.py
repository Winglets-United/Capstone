from openpyxl import load_workbook
from Lib import panel as pn

class ExcelParser(object):
    def __init__(self, inFileLoc, outFileLoc):
        self.wb = load_workbook(inFileLoc)
        self.wsNames = self.wb.get_sheet_names()
        self.numOfPans = len(self.wsNames)
        self.ws = self.wb.get_sheet_by_name(self.wsNames[0])  # initialize to first worksheet
        self.rows = [[]]
        self.inputFile = open(outFileLoc, "w")
        self.row_str = []
        self.mainComRow = ["*ISOLV    LAX       LAY       REXPAR    HAG       FLOATX    FLOATY    ITRMAX\n",
                           "*NMACH  MACH\n", "*NALPHA ALPHA\n",
                           "*LATRL    PSI       PITCHQ    ROLLQ     YAWQ      VINF\n",
                           "*NPAN     SREF      CBAR      XBAR      ZBAR      WSPAN\n"]
        self.panComRow = ["*  X1        Y1        Z1        CORD1     COMMENT: ",
                          "*  NVOR      RNCV      SPC      PDL\n",
                          "*  AINC1     AINC2     ITS      NAP        IQUANT    ISYNT     NPP\n",
                          ]
        self.NAP = int(0)
        self.panels = []

    # Column Deliniation spacing (10 spaces for each item)
    def colDel(self, row, size):
        row_str = ""
        for el in row:
            row_str += str(el)
            spaceNum = size - len(str(el))
            for counter in range(0, spaceNum):
                row_str += " "
        row_str += "\n"
        return row_str

    # Reverse column deliniation
    def revColDel(self, row, size):
        row_str = ""
        for el in row:
            spaceNum = size - len(str(el))
            for counter in range(1, spaceNum):
                row_str += " "
            row_str += str(el)
        row_str += "\n"
        return row_str

    # "", except with a specified range of elements accessed
    def revColDel_lim(self, row, size, elStart, elEnd):
        row_str = ""
        for el in range(elStart, elEnd):
            spaceNum = size - len(str(row[el]))
            for counter in range(0, spaceNum):
                row_str += " "
            row_str += str(row[el])
        # row_str += "\n"
        return row_str

    # Column for first, space for everything else
    def col2SpaceDel(self, row, inSize):
        init = 1
        row_str = ""
        for el in row:
            row_str += str(el)
            spaceNum = inSize - len(str(el))
            if init:
                for counter in range(0, spaceNum):
                    row_str += " "
                init = 0
            else:
                row_str += " "
        row_str += "\n"
        return row_str

    # Numbers cascade vertically not horizontally
    def vertDel(self, rows, size):
        row_str = ""
        for el in rows:
            row_str += str(el)
            spaceNum = size - len(str(el))
            for counter in range(0, spaceNum):
                row_str += " "
            row_str += "\n"
        return row_str

    def parseMain(self):
        for col in self.ws.iter_cols(min_row=3, max_col=8, max_row=3):
            for cell in col:
                self.rows[0].append(cell.value)

        self.NMACH = int(self.ws['A5'].value)
        self.rows.append([])
        for col in self.ws.iter_cols(min_row=5, max_col=self.NMACH + 1, max_row=5):
            for cell in col:
                self.rows[1].append(cell.value)

        self.NALPHA = int(self.ws['A7'].value)
        self.PSI = int(self.ws['B9'].value)
        self.rows.append([])
        for col in self.ws.iter_cols(min_row=7, max_col=self.NALPHA + 1, max_row=7):
            for cell in col:
                self.rows[2].append(cell.value)

        self.rows.append([])
        for col in self.ws.iter_cols(min_row=9, max_col=6, max_row=9):
            for cell in col:
                self.rows[3].append(cell.value)

        self.rows.append([])
        for col in self.ws.iter_cols(min_row=11, max_col=6, max_row=11):
            for cell in col:
                self.rows[4].append(cell.value)

        self.row_str.append(self.colDel(self.rows[0], 10))
        self.row_str.append(self.col2SpaceDel(self.rows[1], 8))
        self.row_str.append(self.col2SpaceDel(self.rows[2], 8))
        self.row_str.append(self.colDel(self.rows[3], 10))
        self.row_str.append(self.colDel(self.rows[4], 10))

        self.inputFile.write(self.ws['A1'].value + "\n*\n")
        iter = 0
        for el in self.row_str:
            self.inputFile.write(self.mainComRow[iter])
            self.inputFile.write(el)
            iter += 1
    def parsePanel(self, panNum):
        ws = self.wb.get_sheet_by_name(self.wsNames[panNum])
        panel = pn.panel(ws)
        self.panels.append(panel)
        for row_str in panel.row_str:
            self.inputFile.write(row_str)
        self.inputFile.write("*-----------------------------------------------------\n")
    def parsePanel_old(self, panNum):
        # switch to requested panel sheet
        self.ws = self.wb.get_sheet_by_name(self.wsNames[panNum])
        self.rows.clear()
        self.row_str.clear()
        aincRow = ""
        self.rows.append([])
        for col in self.ws.iter_cols(min_row=2, max_col=4, max_row=2):
            for cell in col:
                self.rows[0].append(cell.value)

        self.rows.append([])
        for col in self.ws.iter_cols(min_row=3, max_col=4, max_row=3):
            for cell in col:
                self.rows[1].append(cell.value)

        self.rows.append([])
        for col in self.ws.iter_cols(min_row=5, max_col=4, max_row=5):
            for cell in col:
                self.rows[2].append(cell.value)

        self.rows.append([])
        for col in self.ws.iter_cols(min_row=7, max_col=7, max_row=7):
            for cell in col:
                self.rows[3].append(cell.value)

                # self.rows.append([])
                # for row in self.ws.iter_cols(min_row=9, max_col=3, max_row=9):
                #  for cell in row:
                #     self.rows[4].append(cell.value)

        self.NAP = int(self.rows[3][3])  # Assign NAP value

        # self.row_str.append(self.revColDel(self.rows[0],10))
        self.row_str.append(self.revColDel(self.rows[0], 10))
        self.row_str.append(self.revColDel(self.rows[1], 10))
        self.row_str.append(self.revColDel(self.rows[2], 10))
        self.row_str.append(self.revColDel(self.rows[3], 10))
        # self.row_str.append(self.colDel(self.rows[4],10))
        if self.NAP != 0:
            self.parseCamber()
        iter = 0
        for el in self.row_str:
            if iter < 1:
                self.inputFile.write(self.panComRow[0] + self.ws['E2'].value + "\n")

            elif iter > 1 and iter < 4:
                self.inputFile.write(self.panComRow[iter - 1])
            self.inputFile.write(el)
            iter += 1
        self.inputFile.write("*-----------------------------------------------------\n")
        return

    def parseCamber(self):
        for col in self.ws.iter_cols(min_row=10, max_col=3, max_row=10 + self.NAP - 1):
            self.rows.append([])
            for cell in col:
                self.rows[len(self.rows) - 1].append(cell.value)
        # rows 4 => X/C
        self.row_str.append("*X/C Percent C\n")
        self.row_str.append(self.vertDel(self.rows[4], 10))
        self.row_str.append("*CAMBER_ROOT\n")
        self.row_str.append(self.vertDel(self.rows[5], 10))
        self.row_str.append("*CAMBER_TIP\n")
        self.row_str.append(self.vertDel(self.rows[6], 10))

    def parseEnd(self):
        self.ws = self.wb.get_sheet_by_name(self.wsNames[0])
        self.rows.clear()
        self.row_str.clear()
        self.rows.append([])
        self.row_str.append("")
        for col in self.ws.iter_cols(min_row=15, max_col=3, max_row=15):
            for cell in col:
                self.rows[0].append(cell.value)
        self.row_str.append("*NXS   NYS   NZS\n")
        self.row_str.append(self.colDel(self.rows[0], 7))
        for i in range(0, len(self.row_str)):
            self.inputFile.write(self.row_str[i])